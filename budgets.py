import imaplib
import email
import os
import re
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
# import vendor category mappings as category_dict
from category_dict import category_dict

detach_dir = "."

# connect email via username & app token
user = os.environ["BUDGETS_EMAIL"]
pwd = os.environ["BUDGETS_PASSWD"]
url = "imap.gmail.com"

m = imaplib.IMAP4_SSL(url)
m.login(user, pwd)
m.select(mailbox="INBOX", readonly=True)
# FOR TESTING: readonly=True

# grab emailids of all unread New Purchases mails
result, emailid = m.search(None, "UnSeen", "ALL")
emailid = emailid[0].split()  # getting emailids in a list

# Open workbook using openpyxl
workbook_name = datetime.now().strftime("%Y") + "-budget_sheet.xlsx"
try:
    wb = load_workbook(filename=workbook_name)
except:
    wb = Workbook()
# New worksheet for
ws = wb.create_sheet(datetime.now().strftime("%b"))
# Remove the annoying default named 'Sheet'
try:
    wb.remove(wb["Sheet"])
except:
    pass

# Worksheet headings
ws.append(["date", "amount", "vendor", "category"])

##########################
##### program starts #####
##########################

for new in emailid:
    # OTHER OPTIONS: RFC822 (UID BODY[TEXT])
    resp, emailbody = m.fetch(new, '(RFC822)')
    my_mail = email.message_from_string(emailbody[0][1].decode())
    emailbody = str(emailbody)

    # date
    date = re.search(
        's*(Sun|Mon|Tue|Wed|Thu|Fri|Sat), \d{1,2} s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{4}', emailbody)
    date = date.group()

    # transaction amount
    amount = re.search('\$\d{0,2}\,{0,1}\d{0,3}\.\d{0,2}', emailbody)
    amount = amount.group()

    transaction_type = email.header.decode_header(
        my_mail['Subject'])  # type of transaction from emailSubject
    transaction_type = str(transaction_type[0])

    ##### func to categorize vendor by type #####
    def f_vendor_category():
        for keys in category_dict:
            if keys in vendor:
                global vendor_found
                vendor_found = (category_dict[keys])
                break
            else:
                vendor_found = ' '

    # new purchases
    if 'New purchase on your credit card' in transaction_type:
        vendor = re.search('(?<=at )(.*?)(?=\.)', emailbody)
        vendor = str(vendor.group())
        f_vendor_category()
        category = vendor_found

    # # recurring purchases
    elif 'New preauthorized payment with your credit card' in transaction_type:
        vendor = re.search(
            '(?<=to =\Sr\Sn)(.*?)(?= on your CIBC Dividend Visa)', emailbody)
        vendor = str(vendor.group())
        f_vendor_category()
        category = vendor_found

    # returns
    elif 'New purchase return on your credit card' in transaction_type:
        vendor = re.search(
            fr'(?<=\{amount} from )(.*?)(?=\=\\r)', emailbody)
        vendor = str(vendor.group())
        amount = '-{}'.format(amount)
        f_vendor_category()
        category = vendor_found

    else:
        vendor = ""
        category = ""

    ws.append([date, amount, vendor, category])
wb.save(filename=workbook_name)
